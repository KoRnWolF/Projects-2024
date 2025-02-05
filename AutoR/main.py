import pandas
import openpyxl
import json
from datetime import datetime
from tkinter import *
from tkinter import filedialog

tk = Tk()
tk.minsize(height=100, width=100)
tk.title("AutoR")
tk.config(padx=50, pady=50)

file_dir = {
    "Excel" : "",
    "json" : ""
}
def open_file_dialog():
    file_path = filedialog.askopenfilename(title="Select the recon xlsl file", filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")])
    if file_path:
        selected_file_label.config(text=f"Selected Recon File: {file_path}")
    file_dir["Excel"] = file_path

def open_json_dialog():
    json_file_path = filedialog.askopenfilename(title="Select the json recon file", filetypes=[("JSON Files", "*.json"), ("All Files", "*.*")])
    if json_file_path:
        selected_json_label.config(text=f"Selected JSON File: {json_file_path}")
    file_dir["json"] = json_file_path

def process_file():
    try:
        with open(file_dir["json"]) as file:#####change so that user selects the json file
            data = json.load(file)#load json into data

        df = pandas.DataFrame(data)#create panda DF
        reg_df = df.sort_values(by=['timestamp'])#Sort from early to last timestamp
        data_list = reg_df.to_dict('records')#convert DF to dict

        x=0
        row = 13
        #column allocation
        date_col = 1
        from_col = 2
        to_col = 3
        odo_start = 4
        odo_end = 5

        wb = openpyxl.open(file_dir["Excel"])#####change so that user selects the file
        sheet = wb["DRIVER'S LOG BOOK"]

        for trip in data_list:

            trip_date = (datetime.fromisoformat(data_list[x]['timestamp']).strftime("%d-%m-%Y"))
            f_site = (data_list[x]['fromSite'])
            t_site = (data_list[x]['toSite'])
            s_odo = data_list[x]['startOdo']
            e_odo = data_list[x]['endOdo']

            sheet.cell(row=row, column=date_col).value = trip_date
            sheet.cell(row=row, column=from_col).value = f_site
            sheet.cell(row=row, column=to_col).value = t_site
            sheet.cell(row=row, column=odo_start).value = s_odo
            sheet.cell(row=row, column=odo_end).value = e_odo

            x += 1
            row += 1

        wb.save(file_dir["Excel"])###change so that user selects file location
        wb.close()

    except Exception as e:
        selected_file_label.config(text=f"Error: {str(e)}")

open_button = Button(text="Open Recon File", width=15,overrelief="solid", command=open_file_dialog)
open_button.grid(row = 0, column = 0)

open_json_button = Button(text="Open JSON File", width=15,overrelief="solid", command=open_json_dialog)
open_json_button.grid(row = 1, column = 0)

process_recon = Button(text="Process Recon", width=15,overrelief="solid", command=process_file)
process_recon.grid(row = 2, column = 0)

description_files = Label(text="Selected File Directories:")
description_files.grid(row = 3, column = 0)
selected_file_label = Label(text="Selected Recon File")
selected_file_label.grid(row = 4, column = 0)
selected_json_label = Label(text="Selected JSON File(From CatBell)")
selected_json_label.grid(row = 5, column = 0)

tk.mainloop()



