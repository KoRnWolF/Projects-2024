import openpyxl
import os
from datetime import datetime


def split_column_g_lines(input_file, output_file=None):
    """
    Split multi-line text in column G into separate rows.

    Args:
        input_file (str): Path to the input Excel file
        output_file (str, optional): Path to save the output Excel file.
                                     If None, creates a file with timestamp.

    Returns:
        dict: Statistics about the operation
    """
    # Create a default output filename if none provided
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base, ext = os.path.splitext(input_file)
        output_file = f"{base}_split_lines_{timestamp}{ext}"

    # Load the workbook and first worksheet
    wb = openpyxl.load_workbook(input_file)
    source_sheet = wb.active

    # Create a new worksheet for the split lines
    dest_sheet = wb.create_sheet("Split Lines")

    # Copy the header row
    for col_idx, cell in enumerate(source_sheet[1], start=1):
        dest_sheet.cell(row=1, column=col_idx, value=cell.value)

    # Initialize counters
    original_row_count = source_sheet.max_row - 1  # Excluding header
    new_row_count = 0
    dest_row = 2  # Start after header

    # Process each row in the source sheet
    for src_row in range(2, source_sheet.max_row + 1):
        # Get the value in column G (which is the 7th column)
        g_value = source_sheet.cell(row=src_row, column=7).value

        # If cell is empty, just copy the row as is
        if g_value is None or g_value == "":
            for col_idx in range(1, source_sheet.max_column + 1):
                dest_sheet.cell(row=dest_row, column=col_idx,
                                value=source_sheet.cell(row=src_row, column=col_idx).value)
            dest_row += 1
            new_row_count += 1
            continue

        # Convert to string and split by line breaks
        g_value = str(g_value)
        lines = g_value.replace('\r\n', '\n').replace('\r', '\n').split('\n')

        # Process each line
        for line in lines:
            line = line.strip()
            if line:  # Skip empty lines
                # Copy all cells from the source row
                for col_idx in range(1, source_sheet.max_column + 1):
                    # For column G, use the split line
                    if col_idx == 7:
                        value = line
                    else:
                        value = source_sheet.cell(row=src_row, column=col_idx).value

                    dest_sheet.cell(row=dest_row, column=col_idx, value=value)

                dest_row += 1
                new_row_count += 1

    # Auto-adjust column widths
    for column in dest_sheet.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        dest_sheet.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    wb.save(output_file)

    return {
        "success": True,
        "original_row_count": original_row_count,
        "new_row_count": new_row_count,
        "output_file": output_file,
        "message": f"Successfully split lines in column G. Original data had {original_row_count} rows. "
                   f"Created {new_row_count} total rows. Saved to {output_file}"
    }


if __name__ == "__main__":
    # For simplicity, just use hardcoded file paths
    input_file = r"E:\pycharm\Projects\linesplitter\Team Info.xlsx"
    output_file = r"E:\pycharm\Projects\linesplitter\Team Info Split.xlsx"

    # Run the function
    result = split_column_g_lines(input_file, output_file)

    # Print the result
    print(result["message"])

# If you want to use command line arguments instead, uncomment the following:
"""
if __name__ == "__main__":
    import argparse

    # Create command line argument parser
    parser = argparse.ArgumentParser(description="Split multi-line text in Excel column G into separate rows")
    parser.add_argument("input_file", help="Path to the input Excel file")
    parser.add_argument("-o", "--output", help="Path to save the output Excel file", default=None)

    # Parse arguments
    args = parser.parse_args()

    # Run the function
    result = split_column_g_lines(args.input_file, args.output)

    # Print the result
    print(result["message"])
"""